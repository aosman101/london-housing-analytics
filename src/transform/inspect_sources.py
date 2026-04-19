"""Optional exploration utility: prints sheet names and head rows for PIPR and
the extracted ASHE workbooks. Not required for the pipeline — kept to document
how the raw files were inspected during development.

Run `python src/transform/inspect_sources.py` to see sheet structure. The
required ASHE zip extraction now lives in normalise_sources.ensure_ashe_extracted.
"""

import sys
import zipfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from src.common import config  # noqa: E402

RAW = config.RAW_DIR


def preview_excel(path: Path, max_rows: int = 8) -> None:
    print(f"\n===== {path.name} =====")
    wb = load_workbook(path, read_only=True, data_only=True)
    print("Sheets:")
    for s in wb.sheetnames:
        print(f"  - {s}")
    wb.close()

    for s in load_workbook(path, read_only=True).sheetnames[:8]:
        print(f"\n--- Preview: {s} ---")
        try:
            sample = pd.read_excel(path, sheet_name=s, header=None, nrows=max_rows)
            print(sample.to_string(index=False, header=False))
        except Exception as e:
            print(f"Preview failed for {s}: {e}")


def list_zip_contents(path: Path) -> None:
    print(f"\n===== {path.name} =====")
    with zipfile.ZipFile(path, "r") as z:
        print("ZIP contents:")
        for m in z.namelist():
            print(f"  - {m}")


if __name__ == "__main__":
    pipr = config.raw_path("pipr_monthly_price_statistics")
    ashe_zip = config.raw_path("ashe_table8")
    ashe_extract_dir = RAW / "ashe_extracted"

    if pipr.exists():
        preview_excel(pipr)

    if ashe_zip.exists():
        list_zip_contents(ashe_zip)

    if ashe_extract_dir.exists():
        for p in ashe_extract_dir.rglob("*"):
            if p.suffix.lower() in (".xlsx", ".xlsm"):
                preview_excel(p)
